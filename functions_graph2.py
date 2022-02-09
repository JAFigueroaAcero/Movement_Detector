# -*- coding: utf-8 -*-
"""
Created on Thu Oct 21 18:39:38 2021

@author: Juan Antonio

Title: funciones de grafica

ver: 2.0
"""


import matplotlib.pyplot as plt
import os
import openpyxl as op
import numpy as np
import statistics as stat

def ema(lista, inter):
    listac = lista.copy()
    SMA = 0
    EMA = list()
    for a in range(inter):
        SMA += listac.pop(0)
        EMA.append(0)
    SMA = SMA/inter
    EMA.pop(0)
    EMA.append(SMA)
    for el in listac:
        loc = el * (2/(1+inter)) + EMA[-1] * (1 -(2/(1+inter)))
        EMA.append(loc)
    return EMA
def pop(n,pos, *listas):
    try:
        if pos == 0:
            for a in range(n):
                for lista in listas:
                    lista.pop(0)
        elif pos == 1:
            for a in range(n):
                for lista in listas:
                    lista.pop(-1)
    except:
        pass

def graph(name,ref,x, *y):
    plt.title(name)
    for el in y:
        plt.plot(x,el)
    plt.savefig(f'{ref}\\{name}.png')
    plt.show()
    
def mm(lista):
    lon = len(lista)
    i = 0
    val = list()
    maximos = True
    ma = list()
    mi = list()
    while i < lon:
        if i == lon-1:
            pass
        else:
            if lista[i] > lista[i+1] and maximos:
                val.append(lista[i])
                ma.append(lista[i])
                maximos = False
            elif lista[i] < lista[i+1] and not maximos:
                val.append(lista[i])
                mi.append(lista[i])
                maximos = True
            else:
                val.append(None)
        i += 1
    return val, ma, mi

def rel(m,rel):
    rel2 = list()
    m2 = list()
    for a in range(len(m)):
        if m[a] != None:
            rel2.append(rel[a])
            m2.append(m[a])
    return rel2, m2

def graphm(name,ref,x,y):
    plt.title(name)
    if len(x) == 1:
        for el in y:
            plt.plot(x[0],el)
    else:
        for a in range(len(x)):
            plt.plot(x[a],y[a])
    plt.savefig(f'{ref}\\{name}.png')
    plt.show()

def pm(x,y):
    prom_x = list()
    prom_y = list()
    for a in range(len(x)-1):
        prom_x.append((x[a]+x[a+1])/2)
        prom_y.append((y[a]+y[a+1])/2)
    return prom_x, prom_y

def gd(lx,ly,ema_r = 0, ema_i = 0,popi = 0, popd = 0, name = 'def',ref ='def'):
    if ref == 'def':
        try:
            os.mkdir('def')
        except:
            pass
    x = lx.copy()
    y = ly.copy()
    ay,ma,mi = mm(y)
    a,b = rel(ay,x)
    xa, ya = pm(a,b)
    for e in range(ema_r):
        try:
            ya = ema(ya, ema_i)
        except:
            pass
    pop(popi,0, xa,ya,x,y)
    pop(popd,1, xa,ya,x,y)
    
    graphm(name,ref,[x,xa],[y,ya])
    return a,b, ma, mi


def dif(lista):
    lon = len(lista)
    lista_d = list()
    for x in range(1,lon):
        lista_d.append(abs(lista[x]- lista[x-1]))
    return lista_d


def excel(lista,name, ref):
    print('Excel abierto')
    wb = op.Workbook()
    hoja = wb.active
    hoja.title = 'data'
    
    lon = len(lista)
    for x in range(lon):
        hoja.cell(row=1, column=x+1, value=lista[x][0])
        
    x = 1
    for el in lista:
        y = 2
        try:
            for e in el[1]:
                hoja.cell(row=y, column=x, value=e)
                y += 1
        except:
            hoja.cell(row=y, column=x, value= el[1])
        x+=1
    wb.save(f'{ref}\\{name}.xlsx')
    print(f'Excel guardado como: {ref}\\{name}.xlsx')

def hist(lista, classn, ruidomax=0, ruidomin=0, name= 'def', ref ='def'):
    if ref == 'def':
        try:
            os.mkdir('def')
        except:
            pass
    listac = lista.copy()
    for n in range(ruidomax):
        listac.pop(listac.index(max(listac)))
    for n in range(ruidomin):
        listac.pop(listac.index(min(listac)))
    
    listac.sort()
    mini = listac[0]
    maxi = listac[-1]
    rango = abs(maxi-mini)
    classel = rango/classn
    bins = round(((maxi - mini) / classel))
    plt.title(name)
    plt.hist(listac, bins = bins)
    plt.savefig(f'{ref}\\{name}.png')
    plt.show()
    
    
def medianaf(lista):
    ind = int(len(lista)/2 - 0.5)
    
    if len(lista) % 2 == 0:
        return (lista[ind]+lista[ind+1])/2
    else:
        return lista[ind]
    
def md(lista,popd=0, popi=0):
    listac= lista.copy()
    listac.sort()
    print(len(listac))
    pop(popd,1,listac)
    pop(popi,0,listac)
    print(len(listac))
    rango = max(listac) - min(listac)
    varianza = np.var(listac)
    desviacion = varianza**(0.5)
    
    media = sum(listac)/len(listac)
    moda = stat.mode(listac)
    mediana = medianaf(listac)
    
    cov = desviacion/media
    
    return rango, varianza, desviacion, cov, media, moda, mediana
    



def razon(lista):
    loc_l = list()
    for n in range(1,len(lista)):
        loc = lista[n-1]/lista[n]
        loc_l.append(loc)
    return loc_l


